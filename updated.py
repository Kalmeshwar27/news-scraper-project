# ==== Import Libraries ====
import feedparser
import re
import nltk
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# AI Libraries
from keybert import KeyBERT
from transformers import pipeline
from sentence_transformers import SentenceTransformer, util

# ==== NLTK Data Downloads ====
nltk.download('punkt')
nltk.download('stopwords')

# ==== Initialize Models ====
kw_model = KeyBERT()
classifier = pipeline("text-classification", model="distilbert-base-uncased-finetuned-ag-news")
summarizer = pipeline("summarization", model="facebook/bart-large-cnn")
embedding_model = SentenceTransformer('all-MiniLM-L6-v2')

# ==== RSS Feed List ====
rss_feeds = [
    "https://timesofindia.indiatimes.com/rssfeedstopstories.cms",
    "https://www.thehindu.com/news/national/feeder/default.rss",
    "https://indianexpress.com/feed/",
    "https://www.financialexpress.com/feed/",
    "https://feeds.feedburner.com/ndtvnews-top-stories",
    "https://www.hindustantimes.com/feeds/rss/india-news/rssfeed.xml",
    "https://www.livemint.com/rss/news",
    "https://www.bloomberg.com/feed/podcast/etf-report.xml",
    "https://www.france24.com/en/rss",
    "http://feeds.bbci.co.uk/news/rss.xml",
    "https://moxie.foxnews.com/feedburner/latest.xml",
    "https://abcnews.go.com/abcnews/topstories",
    "https://feeds.skynews.com/feeds/rss/home.xml",
    "https://www.aljazeera.com/xml/rss/all.xml",
    "https://www.theguardian.com/world/rss",
    "https://rss.nytimes.com/services/xml/rss/nyt/HomePage.xml",
]

# ==== Containers ====
all_articles = []
article_texts = []

# ==== Fetch & Process Articles ====
for feed_url in rss_feeds:
    feed = feedparser.parse(feed_url)
    for entry in feed.entries:
        title = entry.get("title", "").strip()
        link = entry.get("link", "").strip()
        pub_date = entry.get("published", entry.get("updated", "N/A")).strip()
        raw_description = entry.get("summary", "")
        clean_description = re.sub(r'<[^>]+>', '', raw_description).strip()

        if not title or not link:
            continue

        combined_text = f"{title}. {clean_description}"

        # Keyword Extraction
        keywords = kw_model.extract_keywords(combined_text, top_n=5)
        top_keywords = [f"{kw[0].capitalize()} [{round(kw[1], 2)}]" for kw in keywords]

        # Text Classification
        classification = classifier(combined_text[:512])[0]
        category = classification['label']

        # Summarization
        summary = summarizer(combined_text[:1024])[0]['summary_text']

        article = {
            "title": title,
            "description": clean_description,
            "link": link,
            "published": pub_date,
            "keywords": top_keywords if top_keywords else ["General"],
            "category": category,
            "summary": summary
        }

        all_articles.append(article)
        article_texts.append(combined_text)

# ==== Compute Semantic Similarity ====
embeddings = embedding_model.encode(article_texts, convert_to_tensor=True)
similarity_matrix = util.pytorch_cos_sim(embeddings, embeddings)

# Add top 3 similar articles for each article
for idx, article in enumerate(all_articles):
    sim_scores = similarity_matrix[idx]
    top_indices = sim_scores.argsort(descending=True)[1:4]  # Exclude self
    similar_titles = [all_articles[i]['title'] for i in top_indices]
    article["similar_articles"] = "; ".join(similar_titles)

# ==== Export to Excel ====
wb = Workbook()
ws = wb.active
ws.title = "News Summary"

headers = [
    "News Website URL", "Title", "Description", "Date of Article",
    "Top Keywords", "Category", "Summary", "Similar Articles"
]
ws.append(headers)

header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
bold_font = Font(bold=True)
for col_num, column_title in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = bold_font

for i, article in enumerate(all_articles, start=2):
    ws.cell(row=i, column=1, value=article["link"])
    ws.cell(row=i, column=2, value=article["title"])
    ws.cell(row=i, column=3, value=article["description"])
    ws.cell(row=i, column=4, value=article["published"])
    ws.cell(row=i, column=5, value=", ".join(article["keywords"]))
    ws.cell(row=i, column=6, value=article["category"])
    ws.cell(row=i, column=7, value=article["summary"])
    ws.cell(row=i, column=8, value=article["similar_articles"])

excel_filename = f"news_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(excel_filename)
print(f"✅ Excel created successfully: '{excel_filename}'")
