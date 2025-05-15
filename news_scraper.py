# ==== Import Libraries ====
import feedparser                         # Used to parse RSS feed URLs
import re                                 # For removing HTML tags from descriptions
import nltk                               # Natural Language Toolkit for tokenizing and tagging
from collections import Counter, defaultdict  # For counting keywords and mapping them to articles
from datetime import datetime             # To add timestamp to the Excel filename
from nltk.corpus import stopwords         # List of English stopwords
from nltk import pos_tag, word_tokenize   # Tokenizer and part-of-speech tagger
from openpyxl import Workbook             # For creating Excel workbooks
from openpyxl.styles import Font, PatternFill  # For styling Excel headers

# ==== NLTK Data Downloads ====
nltk.download('punkt')                    # Download tokenizer
nltk.download('averaged_perceptron_tagger')  # Download POS tagger
nltk.download('stopwords')                # Download list of stopwords

# ==== Define English Stopwords and Custom Blacklist ====
stop_words = set(stopwords.words('english'))  # Built-in English stopwords
blacklist = set(stop_words).union({           # Add custom frequent but irrelevant words
    "said", "will", "news", "india", "today", "report", "year", "week", "also",  "time",
    "one", "two", "many", "more", "from", "about",  "could",
     "back", "out", "into", "under", "over","minister", "government",
    "officials", "party", "member", "states", "people", "country", "nation", "issue", "media",
    "world", "video", "audio", "language", "chilli", "pope", "explore", "powder",
    "every", "month", "daily", "newsroom", "click", "read", "update", "headline", "live"
})

# ==== RSS Feed List (Added Top 5 Indian News Sources) ====
rss_feeds = [
    # Indian News Sources (Top 5)
    "https://timesofindia.indiatimes.com/rssfeedstopstories.cms",    # Times of India
    "https://www.thehindu.com/news/national/feeder/default.rss",     # The Hindu
    "https://indianexpress.com/feed/",                                # Indian Express
    "https://www.financialexpress.com/feed/",                         # Financial Express
    "https://feeds.feedburner.com/ndtvnews-top-stories",
    "https://www.hindustantimes.com/feeds/rss/india-news/rssfeed.xml"
    "https://www.livemint.com/rss/news"                               # Livemint
     # Global News Sources
    "https://www.bloomberg.com/feed/podcast/etf-report.xml",
    "https://www.france24.com/en/rss",
    "http://feeds.bbci.co.uk/news/rss.xml",
    "https://moxie.foxnews.com/feedburner/latest.xml",
    "https://abcnews.go.com/abcnews/topstories",
    "https://feeds.skynews.com/feeds/rss/home.xml",
    "https://www.aljazeera.com/xml/rss/all.xml",
    "https://www.theguardian.com/world/rss",
    "https://rss.nytimes.com/services/xml/rss/nyt/HomePage.xml",
]

# ==== Define Category Keywords for Auto-Tagging ====
category_keywords = {
    "Business": {"market", "stocks", "investor", "investment", "trade", "finance", "economy"},
    "Politics": {"election", "president", "policy", "minister", "government", "parliament"},
    "Sports": {"match", "tournament", "player", "goal", "score", "team", "cricket", "football"},
    "Technology": {"ai", "tech", "robot", "software", "hardware", "startup", "gadget"},
    "Climate": {"weather", "climate", "heat", "rain", "storm", "hurricane", "pollution"}
}

# ==== Function to Assign a Category Based on Nouns ====
def assign_category(nouns):
    for category, keywords in category_keywords.items():
        if any(noun in keywords for noun in nouns):  # If any noun matches category keywords
            return category
    return "General"  # Return "General" if no category matched

# ==== Containers to Store Results ====
all_articles = []                         # List to store final processed article dictionaries
keyword_to_articles = defaultdict(list)   # Mapping from keyword to article index for comparison

# ==== Step 1: Fetch & Process Each Article from RSS Feeds ====
for feed_url in rss_feeds:
    feed = feedparser.parse(feed_url)     # Parse RSS feed XML
    
    for entry in feed.entries:
        # Extract basic metadata
        title = entry.get("title", "").strip()
        link = entry.get("link", "").strip()
        pub_date = entry.get("published", entry.get("updated", "N/A")).strip()
        raw_description = entry.get("summary", "")
        clean_description = re.sub(r'<[^>]+>', '', raw_description).strip()  # Remove HTML tags

        if not title or not link:         # Skip incomplete articles
            continue

        # Combine title and description for NLP
        combined_text = f"{title}. {clean_description}"
        tokens = word_tokenize(combined_text)            # Tokenize text
        tagged_words = pos_tag(tokens)                   # Apply POS tagging

        # Extract and count nouns (important keywords)
        nouns = [
            word.lower() for word, tag in tagged_words
            if tag.startswith('NN') and word.isalpha() and len(word) > 3 and word.lower() not in blacklist
        ]
        noun_counts = Counter(nouns)
        top_keywords = [f"{word.capitalize()} [{count}]" for word, count in noun_counts.most_common(5)]

        # Extract and count verbs (for action summarization)
        verbs = [
            word.lower() for word, tag in tagged_words
            if tag.startswith('VB') and word.isalpha() and word.lower() not in blacklist
        ]
        verb_counts = Counter(verbs)
        top_noun = noun_counts.most_common(1)[0][0] if noun_counts else "General"
        top_verb_summary = ", ".join(f"{v} ({c})" for v, c in verb_counts.most_common(3)) if verb_counts else "None"
        topic_with_verbs = f"{top_noun.capitalize()}: {top_verb_summary}"

        # Assign a category based on extracted nouns
        category = assign_category(nouns)

        # Construct the final article dictionary
        article = {
            "title": title,
            "description": clean_description,
            "link": link,
            "published": pub_date,
            "keywords": top_keywords if top_keywords else ["General"],
            "topic_verbs": topic_with_verbs,
            "category": category
        }

        all_articles.append(article)  # Add to final list

        # Map each keyword (without count) to the article index
        for kw in top_keywords:
            keyword_clean = re.sub(r'\s*\[\d+\]', '', kw).lower()
            keyword_to_articles[keyword_clean].append(len(all_articles) - 1)

# ==== Step 2: Find Keywords Shared with Other Articles ====
for i, article in enumerate(all_articles):
    common_keywords = set()
    for kw in article["keywords"]:
        keyword_clean = re.sub(r'\s*\[\d+\]', '', kw).lower()
        if len(keyword_to_articles[keyword_clean]) > 1:
            common_keywords.add(keyword_clean)
    article["common_match"] = ", ".join(sorted(common_keywords)) if common_keywords else ""

# ==== Step 3: Find Common Keywords in Groups of 2 Articles ====
for i in range(0, len(all_articles), 2):  # Group every 2 articles
    group = all_articles[i:i+2]          # Get current group
    keyword_sets = []
    for art in group:
        # Extract just the keyword text (without counts)
        keywords = {re.sub(r'\s*\[\d+\]', '', kw).lower() for kw in art['keywords']}
        keyword_sets.append(keywords)

    # Find common keywords in this group
    if len(keyword_sets) >= 2:
        common = set.intersection(*keyword_sets)
    else:
        common = set()

    group_kw = ", ".join(sorted(common)) if common else ""

    # Store the common keywords back into each article in the group
    for j in range(len(group)):
        all_articles[i + j]["group_common"] = group_kw

# ==== Step 4: Export the Processed Data to Excel File ====
wb = Workbook()                # Create a new Excel workbook
ws = wb.active                 # Get active worksheet
ws.title = "News Summary"     # Set sheet name

# Define headers for columns
headers = [
    "News Website URL", "Title", "Header", "Date of Article",
    "Top Keywords", "Topic + Verbs", "Category", "Common Keywords", "Common Keywords in 2 Articles"
]
ws.append(headers)            # Write headers to the first row

# Style the header cells
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue fill
bold_font = Font(bold=True)   # Bold font
for col_num, column_title in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = bold_font

# Write article data to rows
for i, article in enumerate(all_articles, start=2):
    ws.cell(row=i, column=1, value=article["link"])
    ws.cell(row=i, column=2, value=article["title"])
    ws.cell(row=i, column=3, value=article["description"])
    ws.cell(row=i, column=4, value=article["published"])
    ws.cell(row=i, column=5, value=", ".join(article["keywords"]))
    ws.cell(row=i, column=6, value=article["topic_verbs"])
    ws.cell(row=i, column=7, value=article["category"])
    ws.cell(row=i, column=8, value=article["common_match"])
    ws.cell(row=i, column=9, value=article.get("group_common", ""))

# Save Excel file with a timestamped filename
excel_filename = f"news_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(excel_filename)
print(f"✅ Excel created successfully: '{excel_filename}'")
